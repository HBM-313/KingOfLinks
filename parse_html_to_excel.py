"""
Parses all hadith HTML files from kingoflinks mirror and exports to Excel.
Run from: C:/xampp/htdocs/AlQanas2/
Output:   hadiths_from_html.xlsx
"""

import os
import re
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl.styles import Font, PatternFill, Alignment

HTML_ROOT = r"C:\xampp\htdocs\AlQanas2"

CATEGORIES = {
    'ImamAli': 'الإمام علي',
    'Mkhalfoon': 'المخالفون',
    'AhlAlBait': 'أهل البيت',
    'Aqydatona': 'عقيدتنا',
    'Bhooth': 'بحوث',
    'Seerah': 'السيرة',
    'SwR': 'صور',
    'ImamHussain': 'الإمام الحسين',
    'ImamMahdi': 'الإمام المهدي',
    'Daeefa': 'الأحاديث الضعيفة',
    'Tjseem': 'التجسيم',
    'NabiMohd': 'النبي محمد',
    'ImamHasan': 'الإمام الحسن',
    'Threef': 'الطريف',
    'Abawalnabi': 'آباء النبي',
    'Abdulmtalib': 'عبد المطلب',
    'Hywan': 'الحيوان',
    'RS': 'RS',
    'Abutalib': 'أبو طالب',
    'RAG': 'RAG',
    'Fatima': 'فاطمة الزهراء',
}


# These are website UI elements — never hadith content
UI_JUNK = ['العودة', 'عدد الروايات', 'عدد :', 'عدد:', 'الجزء', 'الصفحة',
           'كتاب ', 'باب :', 'رقم الصفحة']


def is_junk(text):
    for phrase in UI_JUNK:
        if phrase in text:
            return True
    return False


def clean(text):
    if not text:
        return ''
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def get_category(rel_path):
    parts = rel_path.replace('\\', '/').split('/')
    return parts[0] if parts else ''


def get_topic_from_page(soup):
    for p in soup.find_all('p', class_='article-intro'):
        spans = p.find_all('span', class_='article-content')
        for span in spans:
            txt = clean(span.get_text(' ', strip=True))
            if len(txt) < 5:
                continue
            if re.match(r'^[\d\s()]+$', txt):
                continue
            if is_junk(txt):
                continue
            return txt
    title = soup.find('title')
    if title:
        return clean(title.get_text())
    return ''


def parse_block(block_html, topic, category, rel_path):
    block = BeautifulSoup(block_html, 'html.parser')
    paras = block.find_all('p', class_='article-intro')
    if not paras:
        return None

    full_text = clean(block.get_text(' ', strip=True))
    if len(full_text) < 30:
        return None

    source_name = ''
    book_name = ''
    volume = ''
    page = ''
    narrator_chain = ''
    hadith_text = ''

    # SOURCE & BOOK: first span.article-content
    first_span = block.find('span', class_='article-content')
    if first_span:
        parent_p = first_span.find_parent('p')
        if parent_p:
            all_spans = parent_p.find_all('span', class_='article-content')
            raw = clean(' '.join(s.get_text(' ', strip=True) for s in all_spans))
        else:
            raw = clean(first_span.get_text(' ', strip=True))
        raw = raw.strip('.,،:-– ')
        if ' - ' in raw:
            pts = raw.split(' - ', 1)
            source_name = clean(pts[0])
            book_name = clean(pts[1])
        else:
            source_name = raw

    # VOLUME & PAGE
    for p in paras:
        txt = clean(p.get_text(' ', strip=True))
        m = re.search(r'الجزء\s*[:(]\s*(\d+)', txt)
        if m:
            volume = m.group(1)
        m2 = re.search(r'(?:رقم\s+)?الصفحة\s*[:(]\s*([\d/]+)', txt)
        if m2:
            page = m2.group(1)

    # HADITH TEXT
    data_el = block.find(id='Data')
    if data_el:
        hadith_text = clean(data_el.get_text(' ', strip=True))
    else:
        candidates = []
        for p in paras:
            txt = clean(p.get_text(' ', strip=True))
            if is_junk(txt):
                continue
            if len(txt) > 40:
                candidates.append(txt)
        if candidates:
            hadith_text = max(candidates, key=len)

    if not hadith_text or len(hadith_text) < 20:
        return None

    # NARRATOR CHAIN
    for p in paras:
        txt = clean(p.get_text(' ', strip=True))
        patterns = [
            r'(?:حدثنا|أخبرنا)\s*:?\s*.+?(?=\s*قال\s*:)',
            r'عن\s+[\u0600-\u06FF\s]+\s*،\s*(?:عن|قال)',
        ]
        for pat in patterns:
            m = re.search(pat, txt)
            if m and len(m.group(0)) > len(narrator_chain):
                narrator_chain = clean(m.group(0))

    return {
        'hadith_text': hadith_text,
        'narrator_chain': narrator_chain,
        'source_name': source_name,
        'book_name': book_name,
        'volume': volume,
        'page': page,
        'category': category,
        'category_ar': CATEGORIES.get(category, category),
        'topic': topic,
        'file_path': rel_path.replace('\\', '/'),
    }


def parse_file(abs_path, rel_path):
    with open(abs_path, encoding='utf-8', errors='replace') as f:
        raw = f.read()

    soup = BeautifulSoup(raw, 'html.parser')
    category = get_category(rel_path)
    topic = get_topic_from_page(soup)
    blocks = re.split(r'<hr\s*/?>', raw, flags=re.IGNORECASE)

    hadiths = []
    for block_html in blocks:
        result = parse_block(block_html, topic, category, rel_path)
        if result:
            hadiths.append(result)
    return hadiths


def find_html_files(root):
    html_files = []
    skip_dirs = {'__pycache__', '.git', 'node_modules'}
    for dirpath, dirnames, filenames in os.walk(root):
        dirnames[:] = [d for d in dirnames if d not in skip_dirs]
        for fn in filenames:
            if fn.lower().endswith(('.html', '.htm')):
                abs_path = os.path.join(dirpath, fn)
                rel_path = os.path.relpath(abs_path, root)
                html_files.append((abs_path, rel_path))
    return html_files


def main():
    print(f"Scanning: {HTML_ROOT}")
    files = find_html_files(HTML_ROOT)
    print(f"Found {len(files)} HTML files\n")

    all_hadiths = []
    for i, (abs_path, rel_path) in enumerate(files):
        try:
            hadiths = parse_file(abs_path, rel_path)
            all_hadiths.extend(hadiths)
            if (i + 1) % 100 == 0:
                print(f"  {i+1}/{len(files)} files - {len(all_hadiths)} hadiths...")
        except Exception as e:
            print(f"  [ERROR] {rel_path}: {e}")

    print(f"\nTotal extracted: {len(all_hadiths)}")

    df = pd.DataFrame(all_hadiths, columns=[
        'hadith_text', 'narrator_chain', 'source_name', 'book_name',
        'volume', 'page', 'category', 'category_ar', 'topic', 'file_path'
    ])
    df.insert(0, 'row_num', range(1, len(df) + 1))

    output = os.path.join(HTML_ROOT, 'hadiths_from_html.xlsx')
    print(f"Writing: {output}")

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Hadiths')
        ws = writer.sheets['Hadiths']

        widths = {
            'A': 8, 'B': 80, 'C': 50, 'D': 30,
            'E': 40, 'F': 8, 'G': 8, 'H': 15,
            'I': 20, 'J': 40, 'K': 45,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', start_color='0D1B14')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.freeze_panes = 'A2'

        for row_num in range(2, len(df) + 2):
            ws.row_dimensions[row_num].height = 60

    print(f"Done! {output}")
    print(f"Total rows: {len(df)}")


if __name__ == '__main__':
    main()
